# excel_and_main.py

import os
import pandas as pd
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from src.config import PDF_DIRECTORY, EXCEL_OUTPUT_PATH, KEYWORDS, initialize_groq_client
from src.processing_utils import preprocess_text, extract_text_pymupdf, extract_text_with_ocr, is_scanned_pdf, search_full_keywords_in_chunks

# -------------------------- EXCEL UTILITIES -------------------------- #

def save_results_to_excel(results: list, excel_path: str) -> None:
    """Save results to Excel with improved readability."""
    df = pd.DataFrame(results)
    df.to_excel(excel_path, index=False)

    workbook = load_workbook(excel_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    for col in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

    workbook.save(excel_path)

# -------------------------- MAIN PROCESS -------------------------- #

def process_pdfs(directory: str, keywords: list, client) -> list:
    """Process PDFs in a directory and extract structured content."""
    all_results = []
    for filename in os.listdir(directory):
        if filename.endswith(".pdf"):
            file_path = os.path.join(directory, filename)
            print(f"Processing: {filename}")

            if is_scanned_pdf(file_path):
                chunks = extract_text_with_ocr(file_path)
            else:
                chunks = extract_text_pymupdf(file_path)

            chunks = preprocess_text(chunks)
            results = search_full_keywords_in_chunks(chunks, keywords, client)
            all_results.extend(results)
    return all_results

if __name__ == "__main__":
    client = initialize_groq_client("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")
    results = process_pdfs(PDF_DIRECTORY, KEYWORDS, client)
    save_results_to_excel(results, EXCEL_OUTPUT_PATH)
    print(f"Processing complete. Results saved to '{EXCEL_OUTPUT_PATH}'.")
