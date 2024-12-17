# Required Libraries
import os
import re
import textwrap
import pandas as pd
import fitz  # PyMuPDF
import pytesseract
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from groq import Groq

# -------------------------- CONFIGURATION -------------------------- #

# Define keywords to search for
KEYWORDS = [
    "SIL", "NACE", "IGC", "UT", "EN10204-3.2", "HARDNESS TESTING", 
    "ISO 5208", "FET", "PMI", "TAT"
]

# Groq API Configuration
GROQ_API_KEY = "your_api_key_here"  # Replace with your API key
MODEL = "llama3-8b-8192"

# Input and Output Paths
PDF_DIRECTORY = "data"  # Replace with the directory containing PDFs
EXCEL_OUTPUT_PATH = "extracted_keywords_with_groq.xlsx"

# -------------------------- INITIALIZATION -------------------------- #

def initialize_groq_client(api_key: str) -> Groq:
    """Initialize the Groq client with the provided API key."""
    return Groq(api_key=api_key)

# -------------------------- TEXT EXTRACTION -------------------------- #

def extract_text_pymupdf(file_path: str) -> list:
    """Extract text from a regular PDF using PyMuPDF."""
    text_chunks = []
    with fitz.open(file_path) as pdf:
        for page in pdf:
            text_chunks.append(page.get_text("blocks"))  
    return [" ".join(block[4] for block in page_blocks) for page_blocks in text_chunks]

def extract_text_with_ocr(file_path: str) -> list:
    """Extract text from scanned PDFs using OCR."""
    images = convert_from_path(file_path, dpi=300)
    text_chunks = []
    for img in images:
        text = pytesseract.image_to_string(img, lang="eng")
        text_chunks.extend(text.split("\n\n"))  # Split into chunks
    return text_chunks

def is_scanned_pdf(file_path: str) -> bool:
    """Determine if a PDF is scanned by checking if it lacks text."""
    with fitz.open(file_path) as pdf:
        return any(page.get_pixmap() for page in pdf if len(page.get_text("blocks")) == 0)

# -------------------------- TEXT PREPROCESSING -------------------------- #

def preprocess_text(chunks: list) -> list:
    """Clean up and preprocess text chunks."""
    clean_chunks = []
    for chunk in chunks:
        chunk = chunk.strip()  
        chunk = " ".join(chunk.split())  
        clean_chunks.append(chunk)
    return clean_chunks

def format_chunk_general(chunk: str) -> str:
    """Format text chunks for better readability."""
    chunk = re.sub(r'(\.|\;|\:)\s+', r'\1\n', chunk)  # Add line breaks after punctuation
    chunk = re.sub(r'\b(IEC \d{4,}|NACE MR\d{4,}|ISO \d{4,})', r'\n  - \1', chunk)  # Bullet points
    formatted_text = "\n".join(textwrap.fill(line, width=80) for line in chunk.split("\n"))
    return formatted_text.strip()

def preprocess_chunks_for_excel(results: list) -> list:
    """Apply formatting to all chunks before saving to Excel."""
    for result in results:
        result['Chunk'] = format_chunk_general(result['Chunk'])
    return results

# -------------------------- KEYWORD SEARCH -------------------------- #

def search_full_keywords_in_chunks(chunks: list, keywords: list) -> list:
    """
    Search for keywords in text chunks and highlight them.
    
    Highlighting is done by enclosing the keyword in ** for bold appearance.
    """
    results = []
    keyword_patterns = [re.compile(rf"\b{re.escape(keyword)}\b", re.IGNORECASE) for keyword in keywords]

    for chunk in chunks:
        highlighted_chunk = chunk  # Start with the original chunk
        for keyword, pattern in zip(keywords, keyword_patterns):
            if pattern.search(chunk):
                # Highlight the keyword by surrounding it with '**' for bold formatting
                highlighted_chunk = pattern.sub(f"**{keyword}**", highlighted_chunk)
                results.append({"Keyword": keyword, "Chunk": highlighted_chunk})
    return results


# -------------------------- GROQ INTEGRATION -------------------------- #

def generate_summary_report(client: Groq, context: str) -> str:
    """Use Groq API to structure text content into readable format."""
    messages = [
        {
            "role": "user",
            "content": (
                f"You are an expert in structuring text into readable formats.\n"
                f"High priority: Do not add any new content. Add line breaks where necessary.\n"
                f"Here is the context: {context}\n"
                f"**Provide the answer in a structured and readable format.**"
            ),
        }
    ]
    chat_completion = client.chat.completions.create(
        messages=messages, model=MODEL, temperature=0.1
    )
    return chat_completion.choices[0].message.content

# -------------------------- EXCEL UTILITIES -------------------------- #

def save_results_to_excel(results: list, excel_path: str) -> None:
    """Save results to Excel with improved readability."""
    df = pd.DataFrame(results)
    df.to_excel(excel_path, index=False)

    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # Apply word wrapping and adjust column widths
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    for col in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

    workbook.save(excel_path)

def highlight_duplicates_in_excel(file_path: str, column: str = "Chunk") -> None:
    """Highlight duplicate entries in Excel."""
    workbook = load_workbook(file_path)
    sheet = workbook.active

    column_idx = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == column:
            column_idx = col[0].column
            break
    if not column_idx:
        raise ValueError(f"Column '{column}' not found.")

    data = [cell.value for cell in sheet[get_column_letter(column_idx)][1:]]
    duplicates = set(x for x in data if data.count(x) > 1)
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for i, cell in enumerate(sheet[get_column_letter(column_idx)][1:], start=2):
        if cell.value in duplicates:
            cell.fill = fill

    workbook.save(file_path)

# -------------------------- MAIN PROCESS -------------------------- #

def process_pdfs(directory: str, keywords: list, client: Groq) -> list:
    """Process PDFs in a directory and extract structured content."""
    all_results = []
    for filename in os.listdir(directory):
        if filename.endswith(".pdf"):
            file_path = os.path.join(directory, filename)
            print(f"Processing: {filename}")

            if is_scanned_pdf(file_path):
                chunks = extract_text_with_ocr(file_path)
            else:
                chunks = extract_text_pymupdf(file_path)

            chunks = preprocess_text(chunks)
            results = search_full_keywords_in_chunks(chunks, keywords)
            all_results.extend(results)
    return all_results

# -------------------------- ENTRY POINT -------------------------- #

if __name__ == "__main__":
    client = initialize_groq_client(GROQ_API_KEY)
    results = process_pdfs(PDF_DIRECTORY, KEYWORDS, client)
    results = preprocess_chunks_for_excel(results)

    save_results_to_excel(results, EXCEL_OUTPUT_PATH)
    highlight_duplicates_in_excel(EXCEL_OUTPUT_PATH)

    print(f"Processing complete. Results saved to '{EXCEL_OUTPUT_PATH}'.")
