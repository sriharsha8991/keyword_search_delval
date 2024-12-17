import os
import re
import pandas as pd
import fitz  # PyMuPDF
import pytesseract
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Define keywords
keywords = ["Solenoid Valve","Positioner","Volume tanks","Ambient condition","Air filter regulator","Local control box","Local control panel","(MAST)","Speed control"]


# Function to extract text using PyMuPDF
def extract_text_pymupdf(file_path):
    text_chunks = []
    with fitz.open(file_path) as pdf:
        for page in pdf:
            text_chunks.append(page.get_text("blocks"))  # Extract text by blocks
    return [" ".join(block[4] for block in page_blocks) for page_blocks in text_chunks]

# Function to handle scanned PDFs (OCR)
def extract_text_with_ocr(file_path):
    images = convert_from_path(file_path, dpi=300)
    text_chunks = []
    for img in images:
        text = pytesseract.image_to_string(img, lang="eng")
        text_chunks.extend(text.split("\n\n"))  # Split into chunks
    return text_chunks

# Function to determine whether a PDF is scanned
def is_scanned_pdf(file_path):
    with fitz.open(file_path) as pdf:
        return any(page.get_pixmap() for page in pdf if len(page.get_text("blocks")) == 0)

# Function to preprocess extracted text
def preprocess_text(chunks):
    clean_chunks = []
    for chunk in chunks:
        chunk = chunk.strip()  # Remove leading/trailing spaces
        chunk = " ".join(chunk.split())  # Normalize whitespace
        clean_chunks.append(chunk)
    return clean_chunks

# Function to search for full-word keywords in chunks
def search_full_keywords_in_chunks(chunks, keywords):
    results = []
    # Prepare regex patterns for each keyword
    keyword_patterns = [re.compile(rf"\b{re.escape(keyword)}\b", re.IGNORECASE) for keyword in keywords]
    
    for chunk in chunks:
        for keyword, pattern in zip(keywords, keyword_patterns):
            if pattern.search(chunk):
                results.append({"Keyword": keyword, "Chunk": chunk})
    return results

# Function to apply duplicate highlighting in Excel
def highlight_duplicates_in_excel(file_path, column="Chunk"):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Find the column index for the specified column
    col_idx = None
    for col in sheet.iter_cols(1, sheet.max_column, 1, 1):
        if col[0].value == column:
            col_idx = col[0].column
            break
    
    if not col_idx:
        raise ValueError(f"Column '{column}' not found in the sheet.")
    
    # Extract data from the specified column
    data = [cell.value for cell in sheet[get_column_letter(col_idx)][1:]]
    
    # Identify duplicate rows
    duplicates = set(x for x in data if data.count(x) > 1)
    
    # Highlight duplicates
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill
    for i, cell in enumerate(sheet[get_column_letter(col_idx)][1:], start=2):  # Start at row 2
        if cell.value in duplicates:
            cell.fill = fill
    
    # Save the updated workbook
    workbook.save(file_path)
# Function to structurize content within chunks
def structurize_chunk_content(chunk):
    structured_content = ""

    # Split chunk into lines for processing
    lines = chunk.split("\n")
    for line in lines:
        # Detect headers (e.g., lines starting with uppercase or ending with ":")
        if re.match(r"^[A-Z\s]+:$", line) or re.match(r"^#{1,6}\s.*", line):
            structured_content += f"**Header:** {line.strip()}\n"

        # Detect bullet points or lists
        elif re.match(r"^[-*•]\s.*", line):
            structured_content += f"- {line.strip()}\n"

        # Detect numbered lists
        elif re.match(r"^\d+\.\s.*", line):
            structured_content += f"{line.strip()}\n"

        # Detect tables or key-value pairs
        elif ":" in line and len(line.split(":")) == 2:
            key, value = map(str.strip, line.split(":", 1))
            structured_content += f"| {key} | {value} |\n"

        # Default paragraph
        else:
            structured_content += f"{line.strip()}\n"

    return structured_content.strip()

# Function to structurize all chunks
def structurize_all_chunks(chunks):
    structured_chunks = []
    for i, chunk in enumerate(chunks, start=1):
        structured_content = structurize_chunk_content(chunk)
        structured_chunks.append(f"### Chunk {i}\n{structured_content}\n---\n")
    return structured_chunks

# Function to save structured chunks to a file
def save_structured_chunks_to_file(structured_chunks, output_file):
    with open(output_file, "w", encoding="utf-8") as file:
        file.write("\n".join(structured_chunks))
    print(f"Structured chunks saved to {output_file}")
# Markdown detection regex
markdown_patterns = [
    r"^#.+",                 # Headings
    r"^\* .+|^- .+|\d+\. .+",  # Lists
    r"\*\*.*?\*\*|__.*?__",  # Bold
    r"\*.*?\*|_.*?_",        # Italics
    r"\[.*?\]\(.*?\)",       # Links
    r"```.*?```",            # Code blocks
    r"`.*?`"                 # Inline code
]
# Main function
def extract_markdown_content(chunks):
    markdown_chunks = []
    for chunk in chunks:
        for pattern in markdown_patterns:
            if re.search(pattern, chunk, re.MULTILINE):
                markdown_chunks.append(chunk)
                break  # Avoid adding the same chunk multiple times
    return markdown_chunks

def process_pdfs(directory, keywords):
    all_results = []
    all_chunks = []  # To store all chunks for structuring
    for filename in os.listdir(directory):
        if filename.endswith(".pdf"):
            file_path = os.path.join(directory, filename)
            print(f"Processing: {filename}")
            if is_scanned_pdf(file_path):
                chunks = extract_text_with_ocr(file_path)
            else:
                chunks = extract_text_pymupdf(file_path)
            
            chunks = preprocess_text(chunks)
            all_chunks.extend(chunks)  # Collect chunks for structuring
            
            markdown_chunks = extract_markdown_content(chunks)
            results = search_full_keywords_in_chunks(markdown_chunks, keywords)
            all_results.extend(results)

    # Structurize all chunks and save to file
    structured_chunks = structurize_all_chunks(all_chunks)
    save_structured_chunks_to_file(structured_chunks, "structured_chunks.txt")

    return all_results

# Directory containing PDF files
pdf_directory = "data"

# Process PDFs and store results
results = process_pdfs(pdf_directory, keywords)

# Save results to Excel
excel_path = "extracted_keywords.xlsx"
df = pd.DataFrame(results)
df.to_excel(excel_path, index=False)

# Highlight duplicates in Excel
highlight_duplicates_in_excel(excel_path)

print(f"Extraction complete. Results saved to '{excel_path}' with duplicates highlighted.")
